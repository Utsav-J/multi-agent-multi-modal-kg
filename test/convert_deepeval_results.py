"""
Convert DeepEval JSON results to Excel/CSV format.

This script reads DeepEval JSON results and converts them into
well-formatted Excel files with multiple sheets for easy analysis.
"""

import json
import sys
import argparse
import csv
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    print("Warning: pandas/openpyxl not available. Will create CSV only.")

# Add project root to path
project_root = Path(__file__).resolve().parent.parent
test_dir = Path(__file__).resolve().parent
sys.path.append(str(project_root))
sys.path.append(str(test_dir))


def load_json_results(json_path: Path) -> dict:
    """Load JSON results file."""
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def create_summary_sheet(summary: dict):
    """Create summary statistics DataFrame."""
    rows = []
    
    for metric_name, stats in summary.get("results_by_metric", {}).items():
        rows.append({
            "Metric": metric_name,
            "Total Evaluations": stats.get("passed", 0) + stats.get("failed", 0),
            "Passed": stats.get("passed", 0),
            "Failed": stats.get("failed", 0),
            "Errors": stats.get("errors", 0),
            "Pass Rate (%)": stats.get("pass_rate", 0),
            "Average Score": stats.get("average_score", None),
            "Min Score": stats.get("min_score", None),
            "Max Score": stats.get("max_score", None),
        })
    
    if HAS_PANDAS:
        return pd.DataFrame(rows)
    return rows


def create_detailed_results_sheet(detailed_results: list):
    """Create detailed results DataFrame with one row per test case."""
    rows = []
    
    for result in detailed_results:
        row = {
            "Test Case Index": result.get("test_case_index", ""),
            "Query": result.get("input", ""),
            "Category": result.get("category", ""),
            "Overall Passed": result.get("overall_passed", False),
        }
        
        # Add metric scores
        metrics = result.get("metrics", {})
        for metric_name, metric_data in metrics.items():
            score = metric_data.get("score")
            success = metric_data.get("success")
            threshold = metric_data.get("threshold")
            eval_time = metric_data.get("evaluation_time_seconds", 0)
            error = metric_data.get("error", False)
            skipped = metric_data.get("skipped", False)
            
            # Format score
            if score is not None:
                score_str = f"{score:.4f}"
            else:
                score_str = "N/A"
            
            # Add columns for each metric
            row[f"{metric_name}_Score"] = score_str
            row[f"{metric_name}_Success"] = "PASS" if success else "FAIL" if success is False else "N/A"
            row[f"{metric_name}_Threshold"] = threshold if threshold is not None else "N/A"
            row[f"{metric_name}_Time_Seconds"] = f"{eval_time:.2f}" if eval_time else "0.00"
            row[f"{metric_name}_Error"] = "Yes" if error else "No"
            row[f"{metric_name}_Skipped"] = "Yes" if skipped else "No"
        
        rows.append(row)
    
    if HAS_PANDAS:
        return pd.DataFrame(rows)
    return rows


def create_metric_breakdown_sheet(detailed_results: list):
    """Create a long-format sheet with one row per test case-metric combination."""
    rows = []
    
    for result in detailed_results:
        test_idx = result.get("test_case_index", "")
        query = result.get("input", "")
        category = result.get("category", "")
        
        metrics = result.get("metrics", {})
        for metric_name, metric_data in metrics.items():
            score = metric_data.get("score")
            success = metric_data.get("success")
            threshold = metric_data.get("threshold")
            eval_time = metric_data.get("evaluation_time_seconds", 0)
            reason = metric_data.get("reason", "")
            error = metric_data.get("error", False)
            skipped = metric_data.get("skipped", False)
            
            rows.append({
                "Test Case Index": test_idx,
                "Query": query,
                "Category": category,
                "Metric": metric_name,
                "Score": score if score is not None else None,
                "Success": success,
                "Threshold": threshold,
                "Evaluation Time (seconds)": eval_time,
                "Reason": reason,
                "Error": error,
                "Skipped": skipped,
            })
    
    if HAS_PANDAS:
        return pd.DataFrame(rows)
    return rows


def create_category_summary_sheet(detailed_results: list):
    """Create summary statistics grouped by category."""
    from collections import defaultdict
    
    category_stats = defaultdict(lambda: {
        "total": 0,
        "passed": 0,
        "failed": 0,
        "metrics": defaultdict(lambda: {"passed": 0, "failed": 0, "errors": 0, "scores": []}),
    })
    
    for result in detailed_results:
        category = result.get("category", "Unknown")
        category_stats[category]["total"] += 1
        
        if result.get("overall_passed", False):
            category_stats[category]["passed"] += 1
        else:
            category_stats[category]["failed"] += 1
        
        metrics = result.get("metrics", {})
        for metric_name, metric_data in metrics.items():
            success = metric_data.get("success")
            score = metric_data.get("score")
            error = metric_data.get("error", False)
            
            if error:
                category_stats[category]["metrics"][metric_name]["errors"] += 1
            elif success:
                category_stats[category]["metrics"][metric_name]["passed"] += 1
            elif success is False:
                category_stats[category]["metrics"][metric_name]["failed"] += 1
            
            if score is not None:
                category_stats[category]["metrics"][metric_name]["scores"].append(score)
    
    rows = []
    for category, stats in sorted(category_stats.items()):
        row = {
            "Category": category,
            "Total Test Cases": stats["total"],
            "Fully Passed": stats["passed"],
            "Partially Failed": stats["failed"],
            "Pass Rate (%)": (stats["passed"] / stats["total"] * 100) if stats["total"] > 0 else 0,
        }
        
        # Add per-metric statistics
        for metric_name, metric_stats in sorted(stats["metrics"].items()):
            total_metric = metric_stats["passed"] + metric_stats["failed"]
            if total_metric > 0:
                row[f"{metric_name}_Pass_Rate"] = (metric_stats["passed"] / total_metric * 100)
                row[f"{metric_name}_Avg_Score"] = (
                    sum(metric_stats["scores"]) / len(metric_stats["scores"])
                    if metric_stats["scores"]
                    else None
                )
            else:
                row[f"{metric_name}_Pass_Rate"] = None
                row[f"{metric_name}_Avg_Score"] = None
        
        rows.append(row)
    
    if HAS_PANDAS:
        return pd.DataFrame(rows)
    return rows


def convert_to_excel(json_path: Path, output_path: Path = None):
    """Convert JSON results to Excel file with multiple sheets."""
    if not HAS_PANDAS:
        print("ERROR: pandas and openpyxl are required for Excel export.")
        print("   Install with: pip install pandas openpyxl")
        print("   Or use --format csv instead")
        return None
    
    print(f"Loading results from {json_path}...")
    data = load_json_results(json_path)
    
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = json_path.parent / f"deepeval_results_{timestamp}.xlsx"
    
    print(f"Creating Excel file: {output_path}")
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1: Summary Statistics
        summary_df = create_summary_sheet(data.get("summary", {}))
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        print(f"  [OK] Created 'Summary' sheet with {len(summary_df)} metrics")
        
        # Sheet 2: Detailed Results (wide format)
        detailed_df = create_detailed_results_sheet(data.get("detailed_results", []))
        detailed_df.to_excel(writer, sheet_name="Detailed Results", index=False)
        print(f"  [OK] Created 'Detailed Results' sheet with {len(detailed_df)} test cases")
        
        # Sheet 3: Metric Breakdown (long format)
        breakdown_df = create_metric_breakdown_sheet(data.get("detailed_results", []))
        breakdown_df.to_excel(writer, sheet_name="Metric Breakdown", index=False)
        print(f"  [OK] Created 'Metric Breakdown' sheet with {len(breakdown_df)} rows")
        
        # Sheet 4: Category Summary
        category_df = create_category_summary_sheet(data.get("detailed_results", []))
        category_df.to_excel(writer, sheet_name="Category Summary", index=False)
        print(f"  [OK] Created 'Category Summary' sheet with {len(category_df)} categories")
    
    # Auto-adjust column widths
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width
        wb.save(output_path)
        print(f"  [OK] Auto-adjusted column widths")
    except Exception as e:
        print(f"  [WARN] Could not auto-adjust column widths: {e}")
    
    print(f"\nSUCCESS: Excel file created successfully: {output_path}")
    return output_path


def convert_to_csv(json_path: Path, output_path: Path = None):
    """Convert JSON results to CSV file (detailed results only)."""
    print(f"Loading results from {json_path}...")
    data = load_json_results(json_path)
    
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = json_path.parent / f"deepeval_results_{timestamp}.csv"
    
    print(f"Creating CSV file: {output_path}")
    
    # Create detailed results
    rows = create_detailed_results_sheet(data.get("detailed_results", []))
    
    if HAS_PANDAS:
        detailed_df = pd.DataFrame(rows)
        detailed_df.to_csv(output_path, index=False, encoding="utf-8")
        print(f"SUCCESS: CSV file created successfully: {output_path}")
        print(f"   Total rows: {len(detailed_df)}")
        print(f"   Total columns: {len(detailed_df.columns)}")
    else:
        # Use standard library CSV
        if not rows:
            print("ERROR: No data to write")
            return None
        
        fieldnames = list(rows[0].keys())
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        
        print(f"SUCCESS: CSV file created successfully: {output_path}")
        print(f"   Total rows: {len(rows)}")
        print(f"   Total columns: {len(fieldnames)}")
    
    return output_path


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Convert DeepEval JSON results to Excel or CSV format"
    )
    parser.add_argument(
        "json_path",
        type=str,
        help="Path to the DeepEval JSON results file",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output file path (optional, auto-generated if not provided)",
    )
    parser.add_argument(
        "--format",
        choices=["excel", "csv", "both"],
        default="excel",
        help="Output format: excel (default), csv, or both",
    )

    args = parser.parse_args()

    json_path = Path(args.json_path)
    if not json_path.exists():
        print(f"ERROR: JSON file not found: {json_path}")
        return

    output_path = Path(args.output) if args.output else None

    try:
        if args.format in ["excel", "both"]:
            convert_to_excel(json_path, output_path)
        
        if args.format in ["csv", "both"]:
            csv_output = output_path.with_suffix(".csv") if output_path else None
            convert_to_csv(json_path, csv_output)
            
    except Exception as e:
        print(f"ERROR: Error converting results: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()

